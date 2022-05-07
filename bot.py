import discord
from discord.ext import commands
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('./data/embed.xlsx')
ws = wb.active
import os

TOKEN = os.environ['TOKEN']
guild_id=os.environ['GUILD']
welcome_channel=os.environ['WELCOME']



intents = discord.Intents.default()
intents.members = True

bot = commands.Bot(command_prefix='$',intents = intents)

      
#load file
for filename in os.listdir('./cmds'):
    if filename.endswith('.py'):
        bot.load_extension(f'cmds.{filename[:-3]}')

@bot.event

async def on_ready():
    activity = discord.Activity(type=discord.ActivityType.listening, name="$help")
    
    await bot.change_presence(status=discord.Status.online, activity=activity)
    print(">>Bot is Online<<")
#Member Join
@bot.event
async def on_member_join(member):
    channel = bot.get_channel(int(welcome_channel))
    guild = bot.get_guild(int(guild_id))
    print(f'{member} join!')
    await channel.send (f'歡迎{member.mention}加入{guild.name},我們歡迎你')
    await member.send (f'歡迎加入{guild.name},我們歡迎你')

if __name__ == '__main__':
    bot.run(TOKEN)