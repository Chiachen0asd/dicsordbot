
import discord
import datetime
from discord.ext import commands
from core.classes import Cog_Extension
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('./data/embed.xlsx')
ws = wb.active
now = datetime.datetime.now()
time_del = datetime.timedelta(hours=8) 
loc_dt =  now - time_del  

class Activity(Cog_Extension): 

    #edit embed
    async def edit_embed(self,ctx,char):
        message_id = int(ws[char+'5'].value)
        guild_id = int(ws[char+'3'].value)
        title = ws[char+'6'].value
        description = ws[char+'7'].value
        author = ws[char+'11'].value
        fi1_n = ws[char+'8'].value
        fi1_t = ws[char+'9'].value
        fi2_n = ws[char+'10'].value
        fi2_t = ws[char+'11'].value
        role = int(ws[char+'14'].value)
        time = (ws[char+'12'].value)

        ws[char+'15'].value = 'done'
        if description == None:description = title
        now = datetime.datetime.now()
        time_del = datetime.timedelta(hours=8) 
        loc_dt =  now - time_del 
        new_embed=discord.Embed(title=title, description=description, color=0xffdd00,timestamp=loc_dt)
        new_embed.set_author(name="easy bot", url="")
        new_embed.set_thumbnail(url="")
        new_embed.set_footer(text=f'by{author}')    
        if fi1_n != None: new_embed.add_field(name=fi1_n, value=fi1_t, inline=True)
        if fi2_n != None: new_embed.add_field(name=fi2_n, value=fi2_t, inline=True)
        new_embed.add_field(name="時間", value=f'{time}', inline=True)
        channel = ctx.channel
        msg = await channel.fetch_message(message_id)
        await msg.edit(embed=new_embed)
        
        guild =  self.bot.get_guild(guild_id)
        roles = guild.get_role(role)
        members = roles.members
        for member in members:
           # user = guild.get_member(member.id)
            if member != guild.get_member(self.bot.user.id):
                await member.send('hi') 

            
    
    @commands.command()
    async def ac_list(self,ctx,dis=0):
        emoji = []
        roles = []
        title = []
        id = []
        author = []
        done = []
        time = []
        index = int(ws['A1'].value)
        for col in range(2,index+1):
            char = get_column_letter(col)
            id.append((ws[char+'1'].value))
            emoji.append(ws[char+'13'].value)
            roles.append(ws[char+'14'].value)
            title.append(ws[char+'6'].value)
            author.append(ws[char+'11'].value)
            time.append(ws[char+'12'].value)
            done.append(ws[char+'15'].value)
        await ctx.send(f'```Id\tTitle\t  Emoji\t\tRoles\t\t\t Author\t\tDate\tSend```')
        if dis == 0:
            for i in range(len(id)):
                await ctx.send(f'```{id[i]}\t{title[i]}\t{emoji[i]}\t{roles[i]}\t{author[i]}\t{time[i]}\t{done[i]}```')
        else:
            await ctx.send(f'```{id[dis]}\t{title[dis]}\t{emoji[dis]}\t{roles[dis]}\t{author[dis]}\t{time[dis]}\t{done[dis]}```')


    @commands.command()
    async def ac_create(self,ctx,args):
        await ctx.channel.purge(limit=1)
        id = args
        print(id)
        index = int(ws['A1'].value)
        ws['A1'].value = index+1
        char = get_column_letter(index+1)
        ws[char+'1'].value = index
        ws[char+'2'].value = id
        wb.save('./data/embed.xlsx')       


    @commands.command()
    async def ac_title(self,ctx,index,args):
        await ctx.channel.purge(limit=1)
        title = args
        char = get_column_letter(int(index)+1)
        ws[char+'6'].value = title
        wb.save('./data/embed.xlsx')
        if ws[char+'15'].value == 'done':
            await self.edit_embed(ctx,char)
            
    

    @commands.command()
    async def ac_des(self,ctx,index,args):
        await ctx.channel.purge(limit=1)
        description = args
        char = get_column_letter(int(index)+1)
        ws[char+'7'].value = description
        wb.save('./data/embed.xlsx')
        if ws[char+'15'].value == 'done':
            await self.edit_embed(ctx,char)
    @commands.command()
    async def ac_field1(self,ctx,index,arg1,args):
        await ctx.channel.purge(limit=1)
        text = args
        subtitle = arg1
        char = get_column_letter(int(index)+1)
        ws[char+'8'].value = subtitle
        ws[char+'9'].value = text
        wb.save('./data/embed.xlsx')
        if ws[char+'15'].value == 'done':
            await self.edit_embed(ctx,char)
    @commands.command()
    async def ac_field2(self,ctx,index,arg1,args):
        await ctx.channel.purge(limit=1)
        text = args
        subtitle = arg1
        char = get_column_letter(int(index)+1)
        ws[char+'10'].value = subtitle
        ws[char+'11'].value = text
        wb.save('./data/embed.xlsx')
        if ws[char+'15'].value == 'done':
            await self.edit_embed(ctx,char)
    @commands.command()
    async def ac_role(self,ctx,index,role=''):
        await ctx.channel.purge(limit=1)
        role = role[3:].rstrip('>')
        char = get_column_letter(int(index)+1)
        ws[char+'14'].value = str(role)
        wb.save('./data/embed.xlsx')
        if ws[char+'15'].value == 'done':
            await self.edit_embed(ctx,char)
    @commands.command()
    async def ac_emoji(self,ctx,index,emoji=':white_check_mark:'):

        await ctx.channel.purge(limit=1)
        char = get_column_letter(int(index)+1)
        ws[char+'13'].value = emoji
        wb.save('./data/embed.xlsx')
        if ws[char+'15'].value == 'done':
            await self.edit_embed(ctx,char)

    @commands.command()
    async def ac_time(self,ctx,index,year=2050,month=12,day=31,hour=23,min=49):
        await ctx.channel.purge(limit=1)
        char = get_column_letter(int(index)+1)
        date = datetime.datetime(year,month,day,hour,min).strftime('%y/%m/%d %H:%M')
        ws[char+'12'].value = date
        wb.save('./data/embed.xlsx')
        if ws[char+'15'].value == 'done':
            await self.edit_embed(ctx,char)

    @commands.command()
    async def ac_embed(self,ctx,index):
        char = get_column_letter(int(index)+1)
        title = ws[char+'6'].value
        description = ws[char+'7'].value
        emoji = ws[char+'13'].value
        author = ctx.author
        fi1_n = ws[char+'8'].value
        fi1_t = ws[char+'9'].value
        fi2_n = ws[char+'10'].value
        fi2_t = ws[char+'11'].value
        time =  ws[char+'12'].value
        ws[char+'15'].value = 'done'
        if description == None:description = title


        embed=discord.Embed(title=title, description=description, color=0xffdd00,timestamp=loc_dt)
        embed.set_author(name='Easy Bot', url="")
        embed.set_thumbnail(url="")
        embed.set_footer(text=f'by{author}')    
        if fi1_n != None: embed.add_field(name=fi1_n, value=fi1_t, inline=True)
        if fi2_n != None: embed.add_field(name=fi2_n, value=fi2_t, inline=True)
        embed.add_field(name="時間", value=f'{time}', inline=True)
        message = await ctx.send(embed=embed)
        await message.add_reaction(emoji)
        print(message)
        message_id = message.id
        guild_id = message.guild.id
        channel_id = message.channel.id
        ws[char+'16'].value = str(message)
        ws[char+'5'].value = str(message_id)
        ws[char+'3'].value = str(guild_id)
        ws[char+'4'].value = str(channel_id)
        ws[char+'11'].value= str(author)
        wb.save('./data/embed.xlsx')
        
    @commands.Cog.listener()
    async def on_raw_reaction_add(self,data):
        message_id = []
        channel_id = []
        guild_id = []
        emoji = []
        roles = []
        title = []
        time = []
        index = int(ws['A1'].value)

        for col in range(2,index+1):
            char = get_column_letter(col)
            message_id.append(int(ws[char+'5'].value))
            channel_id.append(int(ws[char+'4'].value))
            guild_id.append(int(ws[char+'3'].value))
            emoji.append(ws[char+'13'].value)
            roles.append(int(ws[char+'14'].value))
            title.append(ws[char+'6'].value)
            time.append(ws[char+'12'].value)
        if len(message_id)!=0:
            id = 0
            for mesid in message_id:
                #print(mesid)
                
                reaction = emoji[id]  
                if mesid == data.message_id and reaction == str(data.emoji):      
                    guild = self.bot.get_guild(data.guild_id)
                    role_id = roles[id]
                    role = guild.get_role(role_id)
                    title = title[id]
                    time = time[id]
                    await data.member.add_roles(role)
                    embed=discord.Embed(title="關注通知", description=f'你關注了`{title}`活動,已提供身分組,活動開始前機器人會通知! \n#若需取消關注再次點擊圖示 {reaction} 即可', color=0x6bff84)
                    embed.add_field(name="活動名稱", value=f'{title}', inline=True)
                    embed.add_field(name="時間", value=f'{time}', inline=True)
                    embed.set_footer(text=f'by {self.bot.user.name}')
                    await data.member.send(embed=embed)
                id += 1 
    @commands.Cog.listener()
    async def on_raw_reaction_remove(self,data):
        
        message_id = []
        emoji = []
        roles = []
        title = []
        index = int(ws['A1'].value)
        for col in range(2,index+1):
            char = get_column_letter(col)
            message_id.append(int(ws[char+'5'].value))
            emoji.append(ws[char+'13'].value)
            roles.append(int(ws[char+'14'].value))
            title.append(ws[char+'6'].value)
        if len(message_id)!=0:
            id = 0
            print(roles[id])
            for mesid in message_id:
                reaction = emoji[id]
                if mesid == data.message_id and reaction == str(data.emoji):      
                    guild = self.bot.get_guild(data.guild_id)
                    role_id = roles[id]
                    role = guild.get_role(role_id)
                    title = title[id]
                    user = guild.get_member(data.user_id)
                    await user.remove_roles(role)
                    embed=discord.Embed(title="關注通知", description=f'你已取消關注`{title}`活動,再次點擊圖示{reaction}即可關注', color=0xffd06b)
                    embed.add_field(name="活動名稱", value=f'{title}', inline=True)
                    embed.set_footer(text=f'by {self.bot.user.name}')
                    await user.send(embed=embed)
                id += 1
               



        
def setup(bot):
    bot.add_cog(Activity(bot))