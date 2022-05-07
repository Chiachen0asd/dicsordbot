import discord
from discord.ext import commands
from core.classes import Cog_Extension
import datetime
import asyncio
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('./data/embed.xlsx')
ws = wb.active
roles = []
time = []
guild_id = []
title = []

class Time(Cog_Extension):
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)
        
        async def pre_task():
            for i in range(int(ws['A1'].value)):
                    #print(i)
                    char = get_column_letter(int(i)+1)
                    #print(char)
                    roles.append(ws[char+'14'].value)
                    time.append(ws[char+'12'].value)
                    guild_id.append(int(ws[char+'3'].value))
                    title.append(ws[char+'6'].value)
        async def time_task():
            await self.bot.wait_until_ready()
            while not self.bot.is_closed():
                
                now = datetime.datetime.now()

                loc_dt = now.strftime('%y/%m/%d %H:%M')
                print(loc_dt)
                
                
                for i in range(len(roles)):
                    if time[i] == loc_dt:
                        guild = self.bot.get_guild(guild_id[i])
                        role = guild.get_role(roles[i])
                        members = role.members
                        title = title[i]
                        time = time[i]
                        embed=discord.Embed(title="關注通知", description=f'`{title}`,即將開始,請進入伺服器一同參與', color=0x6bff84)
                        embed.add_field(name="活動名稱", value=f'{title}', inline=True)
                        embed.add_field(name="時間", value=f'{time}', inline=True)
                        embed.set_footer(text=f'by {self.bot.user.name}')
                        for member in members:
                            if member != guild.get_member(self.bot.user.id):
                                await member.send(embed = embed)
        self.bg_task = self.bot.loop.create_task(pre_task())
        self.bg_task = self.bot.loop.create_task(time_task())
        
def setup(bot):
    bot.add_cog(Time(bot))